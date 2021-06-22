from openpyxl import *
from tkinter import filedialog
from tkinter import *
from datetime import *
import pyexcel as p
import os


#MTBF form converter
def MTBF_conv(files):
    boms = []  
    for bb in files:
        boms.append(bb.get()) #files is dic?
    p.save_book_as(file_name=boms[0], dest_file_name='tmp.xlsx')
    pn = boms[1]
    name = boms[2]
    wb = load_workbook('tmp.xlsx')
    wb.save('tmp.xlsx')
    wb = load_workbook('tmp.xlsx')
    ws = wb.active
    num_cols = [11, 10, 9, 8, 6, 5, 3, 2, 1]
    for x in num_cols:
        ws.delete_cols(x)
    num_rows = [5, 3, 2, 1]
    for x in num_rows:
        ws.delete_rows(x)
    print(ws.max_row)
    ws['A1'] = '名称' 
    ws['B1'] = '零件编号'
    ws['C1'] = '数量'
    ws['D1'] = '位号'
    ws['E1'] = '父装配'
    ws['F1'] = '零件'
    ws.insert_rows(2)
    ws['A2'] = name
    ws['B2'] = pn
    ws['C2'] = 1
    ws['E2'] = '系统'
    ws['F2'] = '0'
    max_row = ws.max_row
    for i in range(3, max_row + 1):
        ws.cell(row=i, column=5).value = 'Interposer'
        ws.cell(row=i, column=6).value = '1'
    datestring = datetime.strftime(datetime.now(), ' %Y-%m-%d_%H_%M_%S') 
    wb.save(pn+ "_"+ datestring+ '.xlsx')
    os.remove('tmp.xlsx')

# def file_modify(file):
#     wb = load_workbook(file)
#     ws = wb.create_sheet()
#     ws = wb.active
#     colA = ws['a']
#     print(len(colA))
#     for i in range(3, len(colA) + 1):
#         ws.cell(row=i, column=5).value = 'Interposer'
#         ws.cell(row=i, column=6).value = '1'
#     wb.save('newfile.xlsx')
        
#choose the file in
def openfile(ent):
    # bom_conv.withdraw()
    file_in = filedialog.askopenfilename(filetypes = (("Excel 97-2003","*.xls"),("all files","*.*")))
    ent.insert(0, file_in) 


def makeform(root, feilds):
    rows = []
    entries = [];
    for ff in feilds:
        row = Frame(root)
        lab = Label(row, width=12, text=ff)
        ent = Entry(row, width=80)
        row.pack(side=TOP, fill=X)
        lab.pack(side=LEFT)
        ent.pack(side=LEFT, fill=X)
        entries.append(ent)
        rows.append(row)
        
    Button(rows[0], text='SEL', command= lambda:openfile(entries[0])).pack(side=RIGHT)
    return entries

def main():
    bom_conv = Tk()
    bom_conv.title('BOM converter')
    BOM_IN = ['BOM_IN', 'Part Number', 'Name']
    BOM = makeform(bom_conv, BOM_IN)
    BOM_CONV = Button(bom_conv, text = 'BOM Convert', fg='#000079', bg='#66B3FF', font=('Arial', 12), 
                      command= lambda:MTBF_conv(BOM)).pack(side = BOTTOM)
    # result = MTBF_conv(BOM)
    # file_add(result)
    bom_conv.mainloop()

if __name__ == "__main__":
    main();
